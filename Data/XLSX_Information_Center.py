import pandas as pd
from openpyxl import Workbook


# Create a new workbook

# Select the active worksheet (first sheet)


class FileLocationHandeling:
    wb = Workbook()

    def __init__(self):
        self.sheet = FileLocationHandeling.wb.active

    def AddElements(self, AppName, PathFile, i):
        self.sheet["A" + str(i)] = AppName
        self.sheet["B" + str(i)] = PathFile

    # Write data to specific cells
    def AddElements(self, AppName, PathFile, Count, i):
        self.sheet["A" + str(i)] = AppName
        self.sheet["B" + str(i)] = PathFile
        self.sheet["C" + str(i)] = Count

    # Save the workbook with the data
    def File_save(self, _fileName):
        FileLocationHandeling.wb.save(_fileName + '.xlsx')

    def removingDuplicates(self, _fileName):
        # Load the Excel file
        wb_reading = Workbook(_fileName + ".xlsx")
        sheet = wb_reading["your_sheet_name"]
        # Get the column data
        data = sheet["B"]
        # Remove duplicates using a set
        unique_data = list(set(data))
        # Write the unique data back to the Excel file (optional)
        sheet["B"] = unique_data
        # Save the Excel file
        wb_reading.save("your_file_updated.xlsx")

    def PD_GetTheValues(self, pd, row):
        try:

            return pd.loc[row + 1, 2]
        except Exception:
            return 0

    def ReadData(self, _fileName):
        # Read the entire Excel file into a DataFrame
        try:
            df = pd.read_excel(_fileName + ".xlsx")
            return df
        except Exception:
            print("NOt Found")
            wb = Workbook()
            sheet = wb.active
            wb.save(_fileName + ".xlsx")
            df = pd.read_excel(_fileName + ".xlsx")
            return df
        return df

#
# o = FileLocationHandeling()
# o2 = FileLocationHandeling()
# o.AddElements("abc", "a/b/c", 1)
# o2.AddElements("abc22", "a2/b2/c2", 2)
# o2.save("asdfghjkl")
# o.save("asdfghjkl")
