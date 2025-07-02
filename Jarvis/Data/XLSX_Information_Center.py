import pandas as pd
from openpyxl import Workbook, load_workbook
import os
from typing import Optional # Import Optional for type hinting

# Path for storing the Excel file
# Correcting DATA_DIR to be relative to this file's location for robustness
try:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) # Should point to Jarvis directory
    DATA_DIR = os.path.join(BASE_DIR, "Data", "Data_Information_Value")
except NameError: # __file__ is not defined (e.g. interactive, testing)
    # Fallback, assuming script is run from repo root or similar context
    DATA_DIR = os.path.join("Jarvis", "Data", "Data_Information_Value")


if not os.path.exists(DATA_DIR):
    try:
        os.makedirs(DATA_DIR)
    except OSError as e:
        # print(f"Error creating directory {DATA_DIR}: {e}. Files may not save correctly.")
        pass # Proceed, but saving might fail

EXCEL_FILE_NAME_TEMPLATE = os.path.join(DATA_DIR, "{}.xlsx")


def ReadData(_fileName: str) -> Optional[pd.DataFrame]:
    """Reads the entire Excel file into a DataFrame."""
    excel_path = EXCEL_FILE_NAME_TEMPLATE.format(_fileName)
    try:
        if not os.path.exists(excel_path):
            # print(f"File '{excel_path}' not found. Creating a new one for {_fileName}.")
            wb = Workbook()
            sheet = wb.active
            if _fileName == "AppNameList": # Add header if creating AppNameList
                sheet["A1"] = "AppName"
                sheet["B1"] = "PathFile"
                sheet["C1"] = "Count"
            wb.save(excel_path)
            # print(f"Created '{excel_path}'.")

        df = pd.read_excel(excel_path)
        return df
    except Exception as e:
        # print(f"Error reading or creating Excel file '{excel_path}' for {_fileName}: {e}")
        # Attempt to create an empty workbook as a last resort if read fails
        try:
            # print(f"Attempting to create a fallback empty Excel file for {_fileName}.")
            wb = Workbook()
            sheet = wb.active
            if _fileName == "AppNameList": # Ensure header in fallback too
                sheet["A1"] = "AppName"
                sheet["B1"] = "PathFile"
                sheet["C1"] = "Count"
            wb.save(excel_path)
            # print(f"Fallback Excel file created for {_fileName}.")
            return pd.DataFrame() # Return empty DataFrame
        except Exception as ex_inner:
            # print(f"Critical error: Could not create or read Excel file '{excel_path}' for {_fileName}: {ex_inner}")
            return None


def PD_GetTheValues(df: pd.DataFrame, row_index: int, column_index: int = 2):
    """Safely gets a value from a DataFrame using iloc (integer-based indexing)."""
    try:
        return df.iloc[row_index, column_index]
    except IndexError:
        # print(f"Error: Row {row_index} or Column {column_index} out of bounds.")
        return 0
    except Exception as e:
        # print(f"Error getting value from DataFrame: {e}")
        return 0


def file_save_workbook(_fileName: str, wb_to_save: Workbook) -> None:
    """Saves the given workbook to the specified file name using the template."""
    excel_path = EXCEL_FILE_NAME_TEMPLATE.format(_fileName)
    try:
        wb_to_save.save(excel_path)
        # print(f"File saved: {excel_path}")
    except Exception as e:
        # print(f"Error saving Excel file '{excel_path}': {e}")
        pass


class FileLocationHandeling:
    def __init__(self, fileName: str = "AppNameList"):
        self.fileName = fileName
        self.excel_path = EXCEL_FILE_NAME_TEMPLATE.format(self.fileName)
        try:
            if os.path.exists(self.excel_path):
                self.wb = load_workbook(self.excel_path)
            else:
                # print(f"Workbook '{self.excel_path}' not found, creating new one.")
                self.wb = Workbook()
            self.sheet = self.wb.active # Get active sheet

            # Ensure header exists if sheet is empty or new for AppNameList
            if self.fileName == "AppNameList" and (self.sheet.max_row == 0 or self.sheet["A1"].value is None) :
                # print(f"Sheet for '{self.fileName}' is empty or headerless. Adding headers.")
                self.sheet["A1"] = "AppName"
                self.sheet["B1"] = "PathFile"
                self.sheet["C1"] = "Count"
                if self.sheet.max_row == 0: # If truly empty, means it's a new sheet
                     pass # print(f"Initialized new sheet for {self.fileName} with headers.")
                else: # Header was missing (A1 was None)
                     pass # print(f"Added missing headers to existing sheet for {self.fileName}.")

        except Exception as e:
            # print(f"Error initializing FileLocationHandeling for '{self.fileName}': {e}. Creating a new workbook as fallback.")
            self.wb = Workbook() # Fallback to a new workbook
            self.sheet = self.wb.active
            if self.fileName == "AppNameList": # Ensure header in fallback
                self.sheet["A1"] = "AppName"
                self.sheet["B1"] = "PathFile"
                self.sheet["C1"] = "Count"
                # print(f"Fallback workbook and sheet created for {self.fileName} with headers.")

    def AddElements(self, AppName: str, PathFile: str, Count: int, i: int) -> None:
        # 'i' is the data entry number, used for 'Count' and to determine row.
        # Data entries should start from row 2 if there's a header.
        actual_row = i + 1 # Assuming 'i' starts from 1 for data entries, so row is i+1

        # Check if headers are present and adjust if needed, though init should handle this.
        if self.fileName == "AppNameList" and self.sheet["A1"].value != "AppName":
            # This case should ideally be caught by __init__
            # print("Warning: Headers missing in AddElements, re-initializing them.")
            self.sheet["A1"] = "AppName"
            self.sheet["B1"] = "PathFile"
            self.sheet["C1"] = "Count"

        try:
            self.sheet["A" + str(actual_row)] = AppName
            self.sheet["B" + str(actual_row)] = PathFile
            self.sheet["C" + str(actual_row)] = Count # This is the 'entry number'
        except Exception as e:
            # print(f"Error adding elements (AppName: {AppName}) to sheet at row {actual_row}: {e}")
            pass

    def save_workbook(self) -> None:
        """Saves the current instance's workbook."""
        file_save_workbook(self.fileName, self.wb)

# Make File_save available for external use if SystemFilePathScanner directly calls it
# This was the original name of the save function.
File_save = file_save_workbook

# Expose ReadData for external use as well, if SystemFilePathScanner needs it before FLH instance
XLSX_ReadData = ReadData
